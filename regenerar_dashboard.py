#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
regenerar_dashboard.py
----------------------
Descarga los Excel de VENTAS y COMPRAS desde Google Drive (server-side, sin CORS),
extrae las filas tal como las entrega XLSX.js en el navegador (salteando la
columna A vacia) e inyecta los datos como JSON dentro del dashboard HTML,
en la variable  const APP_DATA = ... ;

Soporta dos campanas: 2526 y 2627. El HTML incluye un selector.

Corre en GitHub Actions (ver .github/workflows/actualizar_dashboard.yml).
"""

import requests
import json
import openpyxl
import os
import datetime
from io import BytesIO

# ─── IDs de Google Drive ────────────────────────────────────────────
ARCHIVOS = {
    "2526": {
        "ventas":  "1qyGcHiDbI9S3eNaLbW6nsFPb211njPPP",
        "compras": "17OCs23CBQ4O6En7efejEgnrKkCd6-ckD",
    },
    "2627": {
        "ventas":  "1jXzje3Rg3reYCPfvZDk06kvBnPBO6WDq",
        "compras": "1B-JKT1VgnnYnEMarUzfKkjdkXTzzLOK1",
    },
}

CAMPANA_DEFAULT = "2526"


def download_from_gdrive(file_id):
    """Descarga el archivo .xlsm original desde Google Drive."""
    session = requests.Session()
    url = f"https://drive.google.com/uc?export=download&id={file_id}"
    r = session.get(url, stream=True, timeout=90)
    r.raise_for_status()

    if 'text/html' in r.headers.get('Content-Type', ''):
        import re
        content = r.text
        token_match = re.search(r'confirm=([0-9A-Za-z_\-]+)', content)
        uuid_match = re.search(r'uuid=([0-9A-Za-z_\-]+)', content)
        if token_match:
            token = token_match.group(1)
            url2 = f"https://drive.google.com/uc?export=download&confirm={token}&id={file_id}"
            r = session.get(url2, stream=True, timeout=90)
        elif uuid_match:
            uuid = uuid_match.group(1)
            url2 = f"https://drive.usercontent.google.com/download?id={file_id}&export=download&confirm=t&uuid={uuid}"
            r = session.get(url2, stream=True, timeout=90)
        else:
            url2 = f"https://drive.usercontent.google.com/download?id={file_id}&export=download&confirm=t"
            r = session.get(url2, stream=True, timeout=90)
        r.raise_for_status()

    content = r.content
    if content[:2] != b'PK':
        raise ValueError(f"El archivo {file_id} no es un Excel valido. Primeros bytes: {content[:20]}")
    print(f"  Descargado: {len(content):,} bytes")
    return BytesIO(content)


def cell_to_js(v):
    """Convierte un valor de celda al tipo que XLSX.js entregaria (raw:true)."""
    if v is None:
        return None
    if isinstance(v, datetime.datetime):
        # XLSX.js con cellDates:true entrega un Date; en JSON usamos ISO
        # El dashboard usa toDate() que acepta strings ISO
        return v.strftime('%Y-%m-%dT%H:%M:%S')
    if isinstance(v, datetime.date):
        return v.strftime('%Y-%m-%dT%H:%M:%S')
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return v
    return str(v)


def sheet_to_rows(ws, max_col=30):
    """
    Devuelve la hoja como lista de filas SIN la columna A vacia,
    replicando exactamente lo que XLSX.js entrega en el navegador
    (sheet_to_json con header:1). Cada fila es una lista de valores.
    """
    rows = []
    for row in ws.iter_rows(values_only=True):
        # Saltear la columna A (indice 0) para alinear con XLSX.js
        trimmed = row[1:max_col+1]
        rows.append([cell_to_js(v) for v in trimmed])
    # Quitar filas totalmente vacias al final
    while rows and all(v is None for v in rows[-1]):
        rows.pop()
    return rows


def procesar_campana(camp, ids):
    print(f"\n=== Campana {camp} ===")
    print("Descargando VENTAS...")
    v_file = download_from_gdrive(ids["ventas"])
    print("Descargando COMPRAS...")
    c_file = download_from_gdrive(ids["compras"])

    print("Leyendo VENTAS (hojas VENTAS + STOCK)...")
    wb_v = openpyxl.load_workbook(v_file, read_only=True, data_only=True)
    ventas_rows = sheet_to_rows(wb_v['VENTAS'], max_col=30)
    stock_rows = sheet_to_rows(wb_v['STOCK'], max_col=25) if 'STOCK' in wb_v.sheetnames else []
    wb_v.close()

    print("Leyendo COMPRAS...")
    wb_c = openpyxl.load_workbook(c_file, read_only=True, data_only=True)
    compras_rows = sheet_to_rows(wb_c['COMPRAS'], max_col=30)
    wb_c.close()

    print(f"  VENTAS: {len(ventas_rows)} filas | STOCK: {len(stock_rows)} filas | COMPRAS: {len(compras_rows)} filas")
    return {
        "ventas": ventas_rows,
        "stock": stock_rows,
        "compras": compras_rows,
    }


def update_html(app_data):
    with open('index.html', encoding='utf-8') as f:
        html = f.read()
    raw = json.dumps(app_data, ensure_ascii=False, separators=(',', ':'))
    marker = 'const APP_DATA='
    s = html.find(marker)
    if s == -1:
        raise RuntimeError("No se encontro 'const APP_DATA=' en index.html")
    s += len(marker)
    e = html.find(';/*END_APP_DATA*/', s)
    if e == -1:
        raise RuntimeError("No se encontro el cierre ';/*END_APP_DATA*/' en index.html")
    html = html[:s] + raw + html[e:]
    with open('index.html', 'w', encoding='utf-8') as f:
        f.write(html)
    print(f"\nindex.html actualizado — {len(html):,} caracteres")


def main():
    app_data = {
        "generado": datetime.datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC'),
        "default": CAMPANA_DEFAULT,
        "campanas": {},
    }
    for camp, ids in ARCHIVOS.items():
        try:
            app_data["campanas"][camp] = procesar_campana(camp, ids)
        except Exception as ex:
            print(f"  ERROR en campana {camp}: {ex}")
            app_data["campanas"][camp] = {"ventas": [], "stock": [], "compras": [], "error": str(ex)}

    update_html(app_data)
    total = {c: len(d.get("ventas", [])) for c, d in app_data["campanas"].items()}
    print(f"Listo. Filas de ventas por campana: {total}")


if __name__ == '__main__':
    main()
